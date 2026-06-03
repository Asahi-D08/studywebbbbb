from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

categories = [
    (
        "动物意象 Animal Imagery",
        "FCE4D6",
        [
            "Little skylark",
            "Skylark",
            "Little songbird",
            "Songbird",
            "Little singing bird",
            "Little squirrel",
            "Squirrel",
        ],
    ),
    (
        "大小 / 形态 Size & Infantilizing",
        "DDEBF7",
        [
            "Little Nora",
            "Little girl",
            "Poor little girl",
            "Child",
            "Little woman",
            "Featherhead",
            "Little featherhead",
        ],
    ),
    (
        "花钱 / 性格 Money & Personality",
        "FFF2CC",
        [
            "Spendthrift",
            "My little spendthrift",
            "Extravagant little person",
            "Wasteful",
            "Scatterbrain",
        ],
    ),
    (
        "控制 / 占有 Possessive & Controlling",
        "E2EFDA",
        [
            "My little Nora",
            "My dear little Nora",
            "My pet",
            "My pretty little pet",
            "My little spendthrift",
        ],
    ),
]

wb = Workbook()
ws = wb.active
ws.title = "Nora Nicknames"

header_font = Font(bold=True, size=12, color="1F2937")
cell_font = Font(size=11)
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, (title, color, items) in enumerate(categories, start=1):
    header_cell = ws.cell(row=1, column=col_idx, value=title)
    header_cell.font = header_font
    header_cell.alignment = center
    header_cell.fill = PatternFill("solid", fgColor=color)
    header_cell.border = border

    for row_idx, item in enumerate(items, start=2):
        cell = ws.cell(row=row_idx, column=col_idx, value=item)
        cell.font = cell_font
        cell.alignment = center
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.border = border

    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 32

ws.row_dimensions[1].height = 38
max_rows = max(len(items) for _, _, items in categories) + 1
for r in range(2, max_rows + 1):
    ws.row_dimensions[r].height = 22

ws.freeze_panes = "A2"

out = "Nora_称呼分类.xlsx"
wb.save(out)
print(f"Saved: {out}")
